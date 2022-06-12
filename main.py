import os

import matplotlib.pyplot as plt
import pandas as pd
import openpyxl

SHOW = False


def load_data() -> pd.DataFrame:
    """
    returns a dataframe that contains _all_ the data.
    Each row has a "Year", "State", "Participants", "Passed", "Failed", "FailedPercent", "AvgGrade", "Grade", "Count"
    """
    pickle_file = "/share/Programming/PycharmProjects/Noteninflation/data/dataframe.pkl"
    if os.path.exists(pickle_file):
        return pd.read_pickle(pickle_file)

    columns = ["Year", "State", "Participants", "Passed", "Failed", "FailedPercent", "AvgGrade", "Grade", "Count"]

    rows = []  # cols in the Excel sheet, but let's move to the dataframe state of mind as soon as possible
    for year in range(2006, 2021 + 1):
        filename = f"/share/Programming/PycharmProjects/Noteninflation/data/Aus_Abiturnoten_{year}.xlsx"
        print(f"Reading Workbook {filename}")
        ws = openpyxl.load_workbook(filename=filename, read_only=True)["Noten"]
        for col in "BCDEFGHIJKLMNOPQ":
            row = [year, ws[col + "5"].value]
            row.extend([cell[0].value for cell in ws[f"{col}6:{col}10"]])
            grade_counts = [int(cell[0].value) for cell in ws[f"{col}12:{col}42"]]
            for i, count in enumerate(grade_counts):
                grade = (i + 10) / 10
                rows.append(row + [grade, count])

    re = pd.DataFrame(rows, columns=columns)
    pd.to_pickle(re, pickle_file)
    return re


def plot_simple(df, kind="line", x="Year", y=None, title="Test", grid="both"):
    df.reset_index().plot(kind=kind, x=x, y=y)
    plt.xlabel(x)
    plt.ylabel(y)
    plt.grid(axis=grid)
    plt.title(title)
    if SHOW:
        plt.show()
    plt.savefig(f"images/plot_simple_{title.lower().replace(' ', '_')}.png", dpi=300)
    plt.clf()


def plot_double(df, kind="line", x="Year", y1=None, y2=None, title="Test"):
    if y1 is None:
        y1 = "AgvGrade"
    if y2 is None:
        y2 = "Participants"

    ax = plt.gca()
    df.reset_index().plot(kind=kind, x=x, y=y1, ax=ax, label=y1)
    df.reset_index().plot(kind=kind, x=x, y=y2, ax=ax, label=y2, secondary_y=True)
    plt.xlabel(x)
    ax.grid()
    plt.title(title)
    plt.tight_layout()
    if SHOW:
        plt.show()
    filename = title.lower().replace(' ', '_').replace('\n', '_')
    plt.savefig(f"images/plot_double_{filename}.png", dpi=300)
    plt.clf()


def plot_state_comparison(df, kind="line", x="Year", y=None, title="Test", normalize=None):
    if y is None:
        y = "AgvGrade"

    ax = plt.gca()
    styles = [
        {"linestyle": "-", "marker": "^"},
        {"linestyle": "-", "marker": "D"},
        {"linestyle": "-", "marker": "o"},
        {"linestyle": "-", "marker": "v"}
    ]
    colors = [
        {"color": "brown"},
        {"color": "blue"},
        {"color": "orange"},
        {"color": "green"}
    ]
    for i, (state, state_df) in enumerate(df.groupby("State")):
        # group data
        state_df = state_df.groupby([x]).mean().reset_index()

        # normalize if required
        if normalize == "local":
            val_max = state_df[y].max()
            val_min = state_df[y].min()
            state_df[y] = (state_df[y] - val_min) / (val_max - val_min)
        elif normalize == "global":
            global_max = df[y].max()
            val_max = state_df[y].max()
            state_df[y] = state_df[y] * global_max / val_max

        # plot the data
        state_df.plot(kind=kind, x=x, y=y, ax=ax, label=state, **styles[i % 4], **colors[i // 4])

    plt.xlabel(x)
    plt.ylabel(y)
    plt.grid()
    plt.title(title)
    plt.tight_layout()
    if SHOW:
        plt.show()
    plt.savefig(f"images/plot_by_state_{title.lower().replace(' ', '_')}.png", dpi=300)
    plt.clf()


if __name__ == '__main__':
    dataframe = load_data()

    plot_simple(dataframe.groupby(["Year"]).sum().reset_index(), kind="bar", x="Year", y="Participants", title="Participants per Year")
    plot_simple(dataframe.groupby(["Grade", "Year"]).sum(), kind="bar", x="Grade", y="Count", title="Grade Count per Year", grid="y")
    plot_simple(dataframe.groupby(["Year"]).mean(), kind="line", x="Year", y="AvgGrade", title="Average Grade per Year")
    plot_simple(dataframe.groupby(["Year"]).mean(), kind="line", x="Year", y="FailedPercent", title="Number of Failed Participants each Year")
    plot_simple(dataframe[dataframe.State == "BY"].groupby(["Year"]).mean(), kind="line", x="Year", y="AvgGrade", title="Average Grade per Year in Bavaria")
    plot_simple(dataframe[dataframe.State == "BY"].groupby(["Year"]).mean(), kind="line", x="Year", y="Participants", title="Participants per Year in Bavaria")
    plot_state_comparison(dataframe, y="Participants", title="Participants per State")
    plot_state_comparison(dataframe, y="Participants", normalize="local", title="Participants per State (normalized with method 1)")
    plot_state_comparison(dataframe, y="Participants", normalize="global", title="Participants per State (normalized with method 2)")
    plot_state_comparison(dataframe, y="AvgGrade", title="Average Grade per Year and State")
    plot_state_comparison(dataframe, y="AvgGrade", normalize="local", title="Average Grade per Year and State (normalized with method 1)")
    plot_state_comparison(dataframe, y="AvgGrade", normalize="global", title="Average Grade per Year and State (normalized with method 1)")
    next_title = "Average Grade and Number of Participants per Year in Lower\nSaxony showing introduction and expiration of shortened curriculum"
    plot_double(dataframe[dataframe.State == "NI"].groupby(["Year"]).mean(), y1="AvgGrade", y2="Participants", title=next_title)
